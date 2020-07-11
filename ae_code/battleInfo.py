import openpyxl
import random

"""
1.实现读取人物数据
2.初始化
3.战斗实现，日志
4.计算回合数


"""

# book = openpyxl.load_workbook('ae数值模版.xlsx')

# activeSheet = book['人物属性']

class battleRole(object):
  """docstring for battleRole"""
  def __init__(self,camp,role_id,name,hp,sp,atk,continue_atk,armor,hit,eva,cri,atk_spd,atk_dis,maxhp):
    super(battleRole, self).__init__()
    self.camp           = camp
    self.role_id        = role_id
    self.name           = name
    self.hp             = hp
    self.sp             = sp
    self.atk            = atk
    self.continue_atk   = continue_atk
    self.armor          = armor
    self.hit            = hit
    self.eva            = eva
    self.cri            = cri
    self.atk_spd        = atk_spd
    self.atk_dis        = atk_dis
    self.maxhp          = maxhp
    self.skill_pool     = skill_pool
  def add_skill(self,skill_id):
  	if skill_id not in self.skill_pool:
  		self.skill_pool.append(skill_id)
  	else:
  		print("{}已经拥有这个技能啦".format(self.name,skill_id))
  def returnArmor(self):

  	return self.armor



class RoleSkill(object):
	"""docstring for RoleSkill"""
	def __init__(self, skill_id,func,funcParam1,funcParam2,funcParam3,desc):
		super(RoleSkill, self).__init__()
		self.skillId       = skillId,
		self.func          = func
		self.funcParam1    = funcParam1
		self.funcParam2	   = funcParam2
		self.funcParam3	   = funcParam3
		self.desc          = desc

class BattleFunc(battleRole):
	"""docstring for BattleFunc"""
	def __init__(self, funcId,funcType,funcParam1,funcParam2,funcParam3):
		super(BattleFunc, self).__init__()
		self.funcId        = funcId
		self.funcType      = funcType
		self.funcParam1    = funcParam1
		self.funcParam2	   = funcParam2
		self.funcParam3	   = funcParam3

	def armor_damage(self):
		percentage = self.funcParam1/10000
		armorValue = self.returnArmor()
		damage     = armorValue * percentage
		return damage



		







def init_for_battlecamp(camp:list,campnum:int):
  camp1 =[]
  for i in camp:
    if i.camp == campnum:
      camp1.append(i)

  return camp1

def create_role(camp:list):
	campRoleNum = len(camp)
	print(campRoleNum)
	for i in range(1,campRoleNum+1):
		name = 'camp'+str(camp[0].camp)+"Role"+str(i)
		locals()['camp'+str(camp[0].camp)+"Role"+str(i)] = camp[i]
	print(camp1Role1,camp2Role1)



def normal_battle(role1_atk,role2_hp):
  role2_hp = role2_hp - role1_atk
  return role2_hp

def cri_battle(role1_atk,role2_hp):
  print("暴击了！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！")
  role2_hp = role2_hp - role1_atk*2
  return role2_hp


def hit_chance(role1_hit,role1_eva):
  hitChance =  role1_hit/(role1_hit+role1_eva)
  return hitChance

def dead_check(role_hp):
  if role_hp > 0:
    return False
  else:
    return True
def camp_total_hp(camp:list)->(int):
	campMaxHP = 0
	for i in camp:
		campMaxHP += i.maxhp
	return campMaxHP	


def battle(role1,role2)->(int):
  #role1 atk role2 def
  random_cri = random.random()
  random_hit = random.random()
  hitChance = hit_chance(role1.hit,role1.eva)
  if hitChance >= random_hit:
    if role1.cri >= random_cri:
      role2.hp = cri_battle(role1.atk,role2.hp)
    else:
      role2.hp = normal_battle(role1.atk,role2.hp)
  else:
    print("未命中____________________________________________________________________________")
  print("防守方hp",role2.hp)
  return role2.hp

def battle_result(battlecamp1,battlecamp2)->(tuple):
	print("初始化对战角色")
	# battlecamp1 = init_for_battlecamp(camp,1)#init 一阵营角色
	# battlecamp2 = init_for_battlecamp(camp,2)#init 2 阵营角色

	camp1Num = len(battlecamp1)
	print("camp1 总人数:",camp1Num)
	camp2Num = len(battlecamp2)  
	print("camp2 总人数:",camp2Num)
	#双方阵营初始化血量
	for num1 in range(camp1Num):
		battlecamp1[num1].hp = battlecamp1[num1].maxhp
		print(battlecamp1[num1].name,battlecamp1[num1].maxhp)
	for num2 in range(camp2Num):
		battlecamp2[num2].hp = battlecamp2[num2].maxhp 
		print(battlecamp2[num2].name,battlecamp2[num2].maxhp)


	#双方计算总hp
	tup1 = ()
	camp1MAXhp = camp_total_hp(battlecamp1)
	print("camp1 maxhp:",camp1MAXhp)
	camp2MAXhp = camp_total_hp(battlecamp2)
	print("camp2 maxhp:",camp2MAXhp)
	break_flag = False
	#开始回合循环
	for round in range(1,50):
		print("第{0}回合***************************************************************".format(round))
		if round % 2 == 1:
			print("atk:camp1")
			for num1 in range(camp1Num):
				camp2Dead = 0
				if dead_check(battlecamp1[num1].hp) is not True:
					hitTime = battlecamp1[num1].hit
					for num2 in range(camp2Num):
						if dead_check(battlecamp2[num2].hp) is not True:
							print("def:camp2")
							print('atk:',battlecamp1[num1].name)
							battlecamp2[num2].hp = battle(battlecamp1[num1],battlecamp2[num2])
							print("def：{0}的hp目前为{1}".format(battlecamp2[num2].name,battlecamp2[num2].hp))
							hitTime = hitTime - 1
							if  hitTime == 0:
								break

						else:
							camp2Dead = camp2Dead + 1
							print("camp2 死亡人数:",camp2Dead)

						if camp2Dead == camp2Num:
							print("获胜：camp1")
							print("回合数",round-2)
							print("获胜方DPR",camp2MAXhp/((round+1)/2))
							tup1 = ("camp1",round,camp2MAXhp/((round+1)/2))
							break_flag = True
						if break_flag == True:
							break
				if break_flag == True:
					break
		if break_flag == True:
			break												
				

		elif round % 2 == 0:
			print('atk:camp2')
			for num2 in range(camp2Num):
				camp1Dead = 0
				if dead_check(battlecamp2[num2].hp) is not True:
					hitTime = battlecamp2[num2].hit
					for num1 in range(camp1Num):
						if dead_check(battlecamp1[num1].hp) is not True:
							print("def:camp1")
							print('atk:',battlecamp2[num2].name)
							battlecamp1[num1].hp = battle(battlecamp2[num2],battlecamp1[num1])
							print("def：{0}的hp经过攻击后目前为{1}".format(battlecamp1[num1].name,battlecamp1[num1].hp))
							hitTime = hitTime - 1
							if  hitTime == 0:
								break
						else:
							camp1Dead += 1
							print("camp1 死亡人数:",camp1Dead)
							if camp1Dead == camp1Num:
								print("获胜：camp2")
								print("回合数",round-2)
								print("获胜方DPR",camp1MAXhp/((round+1)/2))
								tup1 = ("camp2",round,camp1MAXhp/((round+1)/2))
								break_flag = True
						if break_flag == True:
							break
				if break_flag == True:
					break
		if break_flag == True:
			break				
      # if dead_check(camp2Role1.hp) is True:
      #   print("获胜：",camp1Role1.name)
      #   print("回合数",round)
      #   print("获胜者DPR",camp2Role1.maxhp/round)
      #   tup1 = (camp1Role1.name,round,camp2Role1.maxhp/round)
      #   break
      # else:
      #   pass

    # if round % 2 == 0:
    #   print("atk",camp2Role1.name)
    #   camp1Role1.hp = battle(camp2Role1,camp1Role1)

    #   if dead_check(camp1Role1.hp) is True:
    #     print("获胜：",camp2Role1.name)
    #     print("回合数",round)
    #     print("获胜者DPR",camp1Role1.maxhp/round)

    #     tup1 = (camp2Role1.name,round,camp1Role1.maxhp/round)
    #     break
    #   else:
    #     pass
	return tup1

def obtain_camp_data()->(list): #获取所有任务数据
	wb = openpyxl.load_workbook('ae数值模板.xlsx')
	sheet = wb['战斗模拟']
	maxRow = sheet.max_row
	maxCol = sheet.max_column
	totalList = []
	for i in range(2,maxRow+1):
		celllist = []
		for j in range(1,maxCol+1):
			cellValue = sheet.cell(row=i, column=j).value
			celllist.append(cellValue)
		totalList.append(celllist)
	return totalList


def dynamically_generate_characters_dict(list:list)->(dict):

#动态生成角色字典，字典value暂时无用
	role_num = len(totalList)
	roles = locals()
	for i in range(1,role_num+1):
		roles['role' + str(i) ] = i
	return roles


totalList = obtain_camp_data()
roles = dynamically_generate_characters_dict(totalList)
role_list =[]
for i in totalList:
	camp        = i[0]
	role_id     = i[1]
	role_name   = i[2]
	role_hp     = i[3]
	role_sp     = i[4]
	role_atk    = i[5]
	role_conatk = i[6]
	role_armor  = i[7]
	role_hit    = i[8]
	role_eva    = i[9]
	role_cri    = i[10]
	role_atkspd = i[11]
	role_atkdis = i[12]
	role_maxhp  = i[13]
	#动态给角色字典赋值
	roles["role"+str(totalList.index(i)+1)] = battleRole(camp,role_id,role_name,role_hp,role_sp,role_atk,role_conatk,role_armor,role_hit,role_eva,role_cri,role_atkspd,role_atkdis,role_maxhp)
	role_list.append(roles["role"+str(totalList.index(i)+1)])
print(role_list[0])
battlecamp1 = init_for_battlecamp(role_list,1)#init 一阵营角色
battlecamp2 = init_for_battlecamp(role_list,2)#init 2 阵营角色
# if __name__ == '__main__':
  # #生成对战双方信息
  # camp1_role = battleRole(1,"道具达人",50,50,6,0,5,1,0.25,0.1,1,1,1,50)
  # camp2_role = battleRole(2,"暗骑士",40,60,8,0,5,1,0.25,0.1,1,1,2,40)
  # camp = [camp1_role,camp2_role]
  # #打开结果文件准备写入
wb = openpyxl.load_workbook('relust.xlsx')
sheet = wb.active
# #写入题头
sheet.cell(row = 1, column = 1, value = "第几次测试" )
sheet.cell(row = 1, column = 2, value = "获胜者")
sheet.cell(row = 1, column = 3, value = "第几回合")
sheet.cell(row = 1, column = 4, value = "DPR")
tup1 =()
for i in range(1,500):
	tup1 = battle_result(battlecamp1,battlecamp2) #进行一次战斗
#写入一行战斗结果
	sheet.cell(row = i+1, column = 1, value = i)
	sheet.cell(row = i+1, column = 2, value = tup1[0] )
	sheet.cell(row = i+1, column = 3, value = tup1[1] )
	sheet.cell(row = i+1, column = 4, value = tup1[2] )
wb.save('relust.xlsx')
#读数据





